import requests
import pandas as pd
import xml.etree.ElementTree as ET
import sys
import os
import argparse
from datetime import datetime
from config.rss_feeds import rss_feeds
from requests.exceptions import SSLError, RequestException
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import json

# Define headers to mimic a request from a web browser
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
}

def save_rss_feeds():
    with open('config/rss_feeds.py', 'w') as f:
        f.write('rss_feeds = ')
        json.dump(rss_feeds, f, indent=4)

# Function to clean up text by removing CDATA tags and trimming whitespace
def clean_text(text):
    if text is None:
        return "N/A"
    return text.strip()

# Helper function to get text from an XML element
def get_element_text(element, default="N/A"):
    return element.text.strip() if element is not None and element.text is not None else default

def fetch_and_write_rss(categories):
    # Create a folder named by the current date
    folder_name = datetime.now().strftime("%Y-%m-%d")
    os.makedirs(folder_name, exist_ok=True)

    for category in categories:
        if category not in rss_feeds:
            print(f"Category '{category}' not found.")
            continue

        # Define the file paths
        csv_file = os.path.join(folder_name, f"{category}_headlines.csv")
        excel_file = os.path.join(folder_name, f"{category}_headlines.xlsx")

        # Create a DataFrame to store the data
        data = []

        # Iterate over each RSS feed URL in the specified category
        for feed in rss_feeds[category]:
            rss_url = feed["url"]
            source = feed["source"]
            try:
                # Fetch the RSS feed
                response = requests.get(rss_url, headers=headers)
                response.raise_for_status()  # Check if the request was successful

                # Parse the RSS feed
                root = ET.fromstring(response.content)

                # Iterate over each item in the RSS feed
                for item in root.findall(".//item"):
                    title = get_element_text(item.find("title"))
                    description = get_element_text(item.find("description"))
                    link = get_element_text(item.find("link"))
                    pub_date = get_element_text(item.find("pubDate"))

                    # Format the link as a hyperlink for Excel
                    hyperlink = f'=HYPERLINK("{link}", "Link")'

                    # Append the item data to the DataFrame
                    data.append([source, title, description, hyperlink, pub_date])

            except SSLError as e:
                print(f"SSL error occurred while fetching {rss_url}: {e}")
            except RequestException as e:
                print(f"Request error occurred while fetching {rss_url}: {e}")
            except ET.ParseError as e:
                print(f"Error parsing the RSS feed from {rss_url}: {e}")

        # Create a DataFrame
        df = pd.DataFrame(data, columns=["Source", "Title", "Description", "Link", "Publication Date"])

        # Write the DataFrame to a CSV file
        df.to_csv(csv_file, index=False)

        # Write the DataFrame to an Excel file
        df.to_excel(excel_file, index=False, engine='openpyxl')

        # Adjust the Excel file for better initial viewing
        wb = load_workbook(excel_file)
        ws = wb.active

        # Set column widths
        column_widths = {
            "A": 20,  # Source
            "B": 30,  # Title
            "C": 50,  # Description
            "D": 15,  # Link
            "E": 20   # Publication Date
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Apply text wrapping to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        wb.save(excel_file)

        print(f"RSS feed data written to {csv_file}")
        print(f"RSS feed data written to {excel_file}")

def add_category(category):
    if category in rss_feeds:
        print(f"Category '{category}' already exists.")
    else:
        rss_feeds[category] = []
        save_rss_feeds()
        print(f"Category '{category}' added.")

def add_feed(category, url, source):
    if category not in rss_feeds:
        print(f"Category '{category}' not found.")
    else:
        rss_feeds[category].append({"url": url, "source": source})
        save_rss_feeds()
        print(f"Feed added to category '{category}'.")

def main():
    parser = argparse.ArgumentParser(description="RSS Feed Aggregator")
    parser.add_argument("categories", nargs="*", help="Categories to fetch RSS feeds for")
    parser.add_argument("--add-category", help="Add a new category")
    parser.add_argument("--add-feed", nargs=3, metavar=("CATEGORY", "URL", "SOURCE"), help="Add a new feed to an existing category")

    args = parser.parse_args()

    if args.add_category:
        add_category(args.add_category)
    elif args.add_feed:
        add_feed(*args.add_feed)
    elif args.categories:
        fetch_and_write_rss(args.categories)
    else:
        parser.print_help()

if __name__ == "__main__":
    main()